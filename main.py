# file: main.py

from flask import Flask, request, jsonify
from openpyxl import load_workbook
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText

app = Flask(__name__)

TEMPLATE_PATH = "EXPREIMB (22)-2.xlsx"

EMAIL_HOST = "smtp.gmail.com"
EMAIL_PORT = 587
EMAIL_ADDRESS = os.environ.get("EMAIL_USER")
EMAIL_PASSWORD = os.environ.get("EMAIL_PASS")
ACCOUNTING_EMAIL = os.environ.get("ACCOUNTING_EMAIL")


@app.route("/submit-expense", methods=["POST"])
def submit_expense():
    data = request.get_json()
    project_number = data.get("project_number")
    entries = data.get("expenses", [])

    try:
        wb = load_workbook(filename=TEMPLATE_PATH)
        ws = wb["Expense Form"]

        # Find first empty row
        row = 7
        while ws.cell(row=row, column=1).value:
            row += 1

        start_row = row  # mark start of inserted rows
        category_col_map = {}
        for col in range(3, 12):  # C to K
            header = ws.cell(row=7, column=col).value
            if header:
                category_col_map[header.strip().lower()] = col

        for entry in entries:
            ws.cell(row=row, column=1, value=entry.get("date"))
            ws.cell(row=row, column=2, value=entry.get("description"))
            cat = entry.get("category", "").strip().lower()
            amt = entry.get("amount")
            if cat in category_col_map:
                ws.cell(row=row, column=category_col_map[cat], value=amt)
            else:
                print(f"[WARN] Unknown category: '{cat}'")
            row += 1

        last_data_row = row - 1
        subtotal_row = row

        # Add Subtotal row
        ws.cell(row=subtotal_row, column=2, value="Subtotal")
        for col in range(3, 12):  # Columns C to K
            col_letter = chr(64 + col)
            formula = f"=SUM({col_letter}7:{col_letter}{last_data_row})"
            ws.cell(row=subtotal_row, column=col, value=formula)

        # Optionally: Add grand total in column L
        ws.cell(row=subtotal_row, column=12, value="=SUM(C{0}:K{0})".format(subtotal_row))
        ws.cell(row=subtotal_row, column=11).value = "TOTAL"

        wb.save(TEMPLATE_PATH)

        send_email_with_attachment(TEMPLATE_PATH, project_number)
        return jsonify({"status": "success"}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


def send_email_with_attachment(filepath, project_number):
    msg = MIMEMultipart()
    msg["From"] = EMAIL_ADDRESS
    msg["To"] = ACCOUNTING_EMAIL
    msg["Subject"] = f"Expense Report for Project {project_number}"
    msg.attach(MIMEText(f"Attached is the expense report for project {project_number}.", "plain"))

    with open(filepath, "rb") as f:
        part = MIMEApplication(f.read(), Name=os.path.basename(filepath))
        part["Content-Disposition"] = f'attachment; filename="{os.path.basename(filepath)}"'
        msg.attach(part)

    with smtplib.SMTP(EMAIL_HOST, EMAIL_PORT) as server:
        server.starttls()
        server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        server.send_message(msg)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=3000)
